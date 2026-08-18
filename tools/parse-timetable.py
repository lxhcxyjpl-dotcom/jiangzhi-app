# 课表解析：2025级临八整合课程周课表 -> 临床1班结构化 JSON
# 用法：python tools/parse-timetable.py <xlsx路径> <输出json>
import openpyxl, json, re, sys
from collections import defaultdict

SRC = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\lxhcx\Desktop\各种作业\2025级临八整合课程周课表-2026-2027第一学期（20260709）(1).xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else r"jiangzhi-app\assets\timetable-data.json"

wb = openpyxl.load_workbook(SRC)  # 保留字体/填充颜色
ws = wb["25临床周课表"]

# 三大块：(首行, 末行, 日期行, 首周号)
BLOCKS = [(6, 17, 4, 1), (22, 33, 20, 7), (37, 48, 35, 13)]
WEEK_COLS = 5  # 周一..周五
DAYS = ["一", "二", "三", "四", "五"]

def period_of(row):
    v = ws.cell(row=row, column=2).value
    if v is None:
        return None
    try:
        return int(v)
    except Exception:
        return None

CLASS_SPEC = re.compile(r"^([1-3](?:、[1-3])*班)(前2/3|后1/3)?$")
INLINE_SPEC = re.compile(r"^(.*?)[\s　]*([1-3](?:、[1-3])*班(?:前2/3|后1/3)?)$")

def classes_of(spec):
    m = re.match(r"^([1-3](?:、[1-3])*)班(前2/3|后1/3)?$", spec)
    if not m:
        return set(), ""
    nums = {int(x) for x in m.group(1).split("、")}
    return nums, (m.group(2) or "")

def parse_cell_text(text, cell_color):
    """把单元格文本拆成条目列表 [{content, forAll, subgroup}]（只保留与1班相关的）"""
    out = []
    lines = [l.strip() for l in str(text).replace("\r", "").split("\n") if l.strip()]
    pending = []
    for ln in lines:
        m = CLASS_SPEC.match(ln)
        if m:
            if pending:
                content = " ".join(pending)
                pending = []
                cls, sub = classes_of(ln)
                if not cls or 1 in cls:
                    out.append({"content": content, "forAll": not cls, "subgroup": sub})
        else:
            m2 = INLINE_SPEC.match(ln)
            if m2 and m2.group(2):
                content, spec = m2.group(1).strip(), m2.group(2)
                cls, sub = classes_of(spec)
                if not cls or 1 in cls:
                    out.append({"content": content, "forAll": not cls, "subgroup": sub})
            else:
                pending.append(ln)
    if pending:
        content = " ".join(pending)
        if re.search(r"[1-3](?:、[1-3])*班", content) is not None:
            # 内容里混有班级标注但没被拆出来，保险起见当作1班相关整条保留
            out.append({"content": content, "forAll": False, "subgroup": ""})
        else:
            out.append({"content": content, "forAll": True, "subgroup": ""})
    return out

def cell_category(cell):
    """黄底=进院见习 红色=肌骨系统 蓝色(theme 4/8)=临床综合技能1"""
    try:
        fg = cell.fill.fgColor
        frgb = getattr(fg, "rgb", None)
        if isinstance(frgb, str) and frgb[-6:].upper() in ("FFFF00", "FFF200", "FFF2CC"):
            return "rounds"
    except Exception:
        pass
    try:
        fc = cell.font.color
        rgb = getattr(fc, "rgb", None)
        theme = getattr(fc, "theme", None)
        if isinstance(rgb, str) and rgb[-6:].upper() in ("C00000", "FF0000"):
            return "musculo"
        if isinstance(theme, int) and theme in (4, 8):
            return "clinical"
    except Exception:
        pass
    return "core"

def parse_date(s, year_anchor=2026):
    m = re.match(r"^(\d{1,2})-(\d{1,2})$", str(s).strip())
    if not m:
        return None
    mm, dd = int(m.group(1)), int(m.group(2))
    y = year_anchor if mm >= 8 else year_anchor + 1
    return f"{y:04d}-{mm:02d}-{dd:02d}"

# 第一步：收集条目 entries[(day,start,end)] -> {weeks:set, texts:list}
entries = defaultdict(list)  # key -> list of {week, content, cat, subgroup}
dates_map = {}  # (week, day) -> date

anchor_map = {}
for rng in ws.merged_cells.ranges:
    anchor_map[(rng.min_row, rng.min_col)] = rng

for first, last, date_row, first_week in BLOCKS:
    for g in range(6):
        week = first_week + g
        col_start = 3 + g * WEEK_COLS
        for d in range(WEEK_COLS):
            col = col_start + d
            dates_map[(week, d + 1)] = parse_date(ws.cell(row=date_row, column=col).value)
    for row in range(first, last + 1):
        p = period_of(row)
        if p is None:
            continue
        for col in range(3, 3 + 6 * WEEK_COLS):
            cell = ws.cell(row=row, column=col)
            v = cell.value
            if not v:
                continue
            rng = anchor_map.get((row, col))
            # 跨行合并 -> 节次跨度
            end = p
            if rng and rng.max_row > row:
                pe = period_of(rng.max_row)
                if pe and pe > p:
                    end = pe
            # 跨列合并 -> 每个被覆盖的 (周, 星期) 都发一条（如十一假期跨 5 列）
            cols = [col]
            if rng and rng.max_col > rng.min_col:
                cols = [c for c in range(rng.min_col, rng.max_col + 1) if 3 <= c <= 32]
            cat = cell_category(cell)
            for c2 in cols:
                g2 = (c2 - 3) // WEEK_COLS
                d2 = (c2 - 3) % WEEK_COLS + 1
                w2 = first_week + g2
                for it in parse_cell_text(v, cat):
                    entries[(d2, p, end)].append({
                        "week": w2, "content": it["content"], "cat": cat,
                        "subgroup": it["subgroup"],
                    })

# 第二步：合并相同 (day,start,end,content,cat) 的 weeks
merged = {}
for (day, start, end), lst in entries.items():
    for it in lst:
        key = (day, start, end, it["content"], it["cat"], it["subgroup"])
        if key not in merged:
            merged[key] = {"day": day, "start": start, "end": end, "name": it["content"],
                           "cat": it["cat"], "subgroup": it["subgroup"], "weeks": set()}
        merged[key]["weeks"].add(it["week"])

courses = []
for key, it in merged.items():
    it["weeks"] = sorted(it["weeks"])
    courses.append(it)
courses.sort(key=lambda c: (c["day"], c["start"], c["name"]))

# 标准节次时间（北大医学部通用作息，8:00 开始）
PERIOD_TIMES = ["08:00-08:50", "09:00-09:50", "10:10-11:00", "11:10-12:00",
                "13:00-13:50", "14:00-14:50", "15:10-16:00", "16:10-17:00",
                "18:00-18:50", "19:00-19:50", "20:00-20:50", "21:00-21:50"]

data = {
    "meta": {
        "title": "临床医学8年制2025级 · 临床1班课表",
        "term": "2026-2027 第一学期",
        "class": "临床1班",
        "weeks": 18,
        "startMonday": "2026-08-31",
        "note": "除特殊标注外均在4教（生化楼），其他教室在逸夫楼；红色=肌骨系统，蓝色=临床综合技能1，黄色=进院见习",
    },
    "periods": [{"n": i + 1, "time": PERIOD_TIMES[i]} for i in range(12)],
    "dates": {f"{w}-{d}": dt for (w, d), dt in sorted(dates_map.items()) if dt},
    "courses": courses,
}
with open(OUT, "w", encoding="utf-8") as fp:
    json.dump(data, fp, ensure_ascii=False, indent=1)
print("OK courses:", len(courses), "->", OUT)
cats = defaultdict(int)
for c in courses:
    cats[c["cat"]] += 1
print("categories:", dict(cats))
names = sorted({c["name"] for c in courses})
print("unique names:", len(names))
